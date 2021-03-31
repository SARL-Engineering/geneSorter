from openpyxl import Workbook
from openpyxl import load_workbook
import csv, os, time

folderName = "C:\\Lab Scripts\\Input"
outputFolder = "C:\\Lab Scripts\\Output"
returnList = []

def extractList(fileName):
    print("extractList")
    inputWB = load_workbook(fileName)
    inputSheet = inputWB.active

    for row in range(2, inputSheet.max_row + 1):
        miniList = []
        for column in range(1, inputSheet.max_column + 1):
            miniList.append(str(inputSheet.cell(row, column).value))

        returnList.append(miniList)
    print("returnList is currently: ")
    print(returnList)


def listToDict(lst):
    print("listToDict")
    global dictionary
    dictionary = {}
    for row in lst:
        dictionary[row[0]] = row

    print("Dictionary is currently: ")
    print(dictionary)

def headerLister(inputName):
    global headerList
    os.chdir(folderName)
    inputWB = load_workbook(inputName)
    inputSheet = inputWB.active
    headerList = []
    for column in range(1, inputSheet.max_column + 1):
        headerList.append(str(inputSheet.cell(1, column).value))


def dictToSheet(dictionary, headerList):
    print("dictToSheet")
    wb = Workbook()
    sheet = wb.active

    for column in range(1, len(headerList) + 1):
        sheet.cell(1, column).value = headerList[column - 1]
    
    row = 2
    for lst in dictionary:
        column = 1
        for item in dictionary[lst]:
            sheet.cell(row, column).value = item

            column += 1
        row += 1

    os.chdir(outputFolder)
    wb.save("Output.xlsx")




os.chdir(folderName)

bookList = []
for inputExcel in os.listdir(folderName):
    bookList.append(inputExcel)
    print("bookList is currently: ")
    print(bookList)

for i in bookList:
    extractList(i)



listToDict(returnList)

headerLister(bookList[0])

dictToSheet(dictionary, headerList)

print("Job Done. Closing in ten seconds")
time.sleep(10)




    
